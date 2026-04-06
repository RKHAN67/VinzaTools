import sys
from pathlib import Path

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def normalize(value):
    if value is None:
        return ""
    text = str(value).strip()
    return text


def sheet_rows(sheet):
    rows = []
    max_cols = 0
    for row in sheet.iter_rows(values_only=True):
        values = [normalize(cell) for cell in row]
        if any(values):
            rows.append(values)
            max_cols = max(max_cols, len(values))

    if not rows:
        return [["(Empty sheet)"]]

    padded = []
    for row in rows:
        padded.append(row + [""] * (max_cols - len(row)))
    return padded


def main():
    input_path, output_path = sys.argv[1:3]
    workbook = load_workbook(filename=input_path, data_only=True)

    doc = SimpleDocTemplate(
        output_path,
        pagesize=landscape(A4),
        leftMargin=12 * mm,
        rightMargin=12 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )

    styles = getSampleStyleSheet()
    story = []

    for index, sheet in enumerate(workbook.worksheets):
        if index > 0:
            story.append(PageBreak())

        story.append(Paragraph(f"{sheet.title}", styles["Title"]))
        story.append(Spacer(1, 6 * mm))

        rows = sheet_rows(sheet)
        col_count = max(len(row) for row in rows)
        usable_width = landscape(A4)[0] - doc.leftMargin - doc.rightMargin
        col_width = usable_width / max(col_count, 1)

        table = Table(rows, repeatRows=1, colWidths=[col_width] * col_count)
        table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e11d48")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ]
            )
        )
        story.append(table)

    doc.build(story)


if __name__ == "__main__":
    main()
